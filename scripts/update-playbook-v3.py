import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY = '041E42'
GOLD = 'B9975B'
CREAM = 'FAF8F7'
WHITE = 'FFFFFF'
TAUPE = 'E4D5D3'
YELLOW = 'FFF3CD'
LIGHT = 'F0EAE9'
RED = 'C0392B'
GREEN = '1A6B3C'
RED_BG = 'FFE0E0'
GREEN_BG = 'E0FFE0'
BLUE_BG = 'E0E8FF'

navy_fill = PatternFill('solid', fgColor=NAVY)
gold_fill = PatternFill('solid', fgColor=GOLD)
cream_fill = PatternFill('solid', fgColor=CREAM)
white_fill = PatternFill('solid', fgColor=WHITE)
yellow_fill = PatternFill('solid', fgColor=YELLOW)
light_fill = PatternFill('solid', fgColor=LIGHT)

navy_font = Font(name='Arial', bold=True, color=WHITE, size=11)
gold_font = Font(name='Arial', bold=True, color=NAVY, size=11)
header_font = Font(name='Arial', bold=True, color=NAVY, size=10)
body_font = Font(name='Arial', color='2A1F28', size=10)
small_font = Font(name='Arial', color='948794', size=9)
fill_font = Font(name='Arial', color='0000FF', size=10)
title_font = Font(name='Arial', bold=True, color=NAVY, size=14)
subtitle_font = Font(name='Arial', bold=True, color=GOLD, size=12)
italic_font = Font(name='Arial', italic=True, color=GOLD, size=9)
red_font = Font(name='Arial', italic=True, color=RED, size=9)

thin_border = Border(
    left=Side(style='thin', color=TAUPE),
    right=Side(style='thin', color=TAUPE),
    top=Side(style='thin', color=TAUPE),
    bottom=Side(style='thin', color=TAUPE),
)

def hdr(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = navy_fill; cell.font = navy_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

def drow(ws, row, cols, alt=False):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = cream_fill if alt else white_fill
        cell.font = body_font; cell.border = thin_border
        cell.alignment = Alignment(vertical='top', wrap_text=True)

try:
    wb = openpyxl.load_workbook('AMP_KPI_Operational_Playbook_v2.xlsx')
    outfile = 'AMP_KPI_Operational_Playbook_v3.xlsx'
except:
    wb = openpyxl.Workbook()
    outfile = 'AMP_KPI_Operational_Playbook_v3.xlsx'

# ═══════════════════════════════════════════════════
# SHEET: Front Office Protocols → KPI Impact
# ═══════════════════════════════════════════════════
if 'Front Office Protocols' in wb.sheetnames:
    del wb['Front Office Protocols']
ws = wb.create_sheet('Front Office Protocols')
ws.sheet_properties.tabColor = '5B8CB9'

r = 1
ws.cell(row=r, column=1, value='Front Office Protocols Mapped to KPIs').font = title_font
r += 1
ws.cell(row=r, column=1, value='Each protocol from the AMP Front Office Playbook linked to the specific KPIs it improves and the expected impact magnitude.').font = small_font
r += 2

# Section A: Patient Acquisition
ws.cell(row=r, column=1, value='A. PATIENT ACQUISITION PROTOCOLS').font = subtitle_font
r += 1
headers = ['Protocol', 'Standard / Script', 'KPIs Impacted', 'Impact Magnitude', 'Measurement', 'YOUR: Current Compliance']
for j, h in enumerate(headers):
    ws.cell(row=r, column=j+1, value=h)
hdr(ws, r, len(headers))
r += 1

acq_data = [
    ['Answer before 3 rings', 'No exceptions. If on another call, acknowledge within 15 sec.', 'New Patient Volume → Revenue', 'HIGH: +391% conversion for <1 min response', 'Call answer time tracking in phone system', ''],
    ['Rule of 3 Names', 'Practice name, your name, caller\'s name in opening', 'Patient Conversion Rate → Unique Patients', 'MEDIUM: Builds rapport, increases booking rate', 'Mystery shop scores', ''],
    ['Credential the 3 P\'s', 'Practice, Provider, Procedure — on every call before offering appointment', 'Consultation Conversion Rate → Revenue per Patient', 'HIGH: Differentiated practices convert 20-30% higher', 'Call monitoring / QA scores', ''],
    ['Price range, not specific price', 'Every treatment is tailored. Position flexibility as a USP.', 'Avg Revenue Per Patient → Revenue', 'MEDIUM: Prevents price anchoring low', 'Call monitoring', ''],
    ['Always offer next step', 'End every call with consultation offer or follow-up action', 'New Patient Volume → Revenue', 'HIGH: Eliminates "dead" inquiries', 'Calls to bookings ratio', ''],
    ['Pre-consultation call', 'Same day or within 48hrs of booking. Review goals, discuss financing.', 'Treatment Acceptance Rate → Rev Per Patient → Revenue', 'VERY HIGH: Higher acceptance rates, more robust treatment plans', 'Pre-call completion rate (target: 100%)', ''],
    ['Speed of digital response', '<1 min = +391%, <30 min = +62%, <24hr = +17% conversion', 'New Patient Volume → Revenue', 'VERY HIGH: Single most important digital conversion factor', 'Average response time by channel', ''],
    ['Capture email/phone on chat', 'Get contact info early in live chat conversations', 'Lead Quality → Conversion Rate', 'MEDIUM: Enables follow-up on dropped chats', 'Chat-to-booking conversion rate', ''],
]

for i, row_data in enumerate(acq_data):
    for j, val in enumerate(row_data):
        ws.cell(row=r, column=j+1, value=val)
    drow(ws, r, len(headers), alt=(i%2==0))
    ws.cell(row=r, column=1).font = header_font
    ws.cell(row=r, column=6).fill = yellow_fill
    ws.cell(row=r, column=6).font = fill_font
    r += 1

r += 2
# Section B: Check-In/Check-Out
ws.cell(row=r, column=1, value='B. CHECK-IN / CHECK-OUT PROTOCOLS').font = subtitle_font
r += 1
for j, h in enumerate(headers):
    ws.cell(row=r, column=j+1, value=h)
hdr(ws, r, len(headers))
r += 1

checkin_data = [
    ['15-second acknowledgment', 'Verbally acknowledge every patient within 15 sec of arrival', 'Patient Satisfaction → Retention → Unique Patients', 'MEDIUM: First impression drives lifetime value', 'Patient survey scores', ''],
    ['Greet by name', 'Review schedule before patients arrive to anticipate names', 'Patient Experience → Rebooking Rate', 'LOW-MEDIUM: Personal touch increases loyalty', 'Staff observation', ''],
    ['Cosmetic Interest Questionnaire', 'Provide CIQ at every check-in — captures concerns beyond booked treatment', 'Avg Revenue Per Patient → Revenue, Injectable %', 'HIGH: Enables provider to offer comprehensive consultation', 'CIQ completion rate', ''],
    ['Proactive delay communication', 'If provider running late, acknowledge before patient asks', 'Patient Satisfaction → Cancel Rate (inverse)', 'MEDIUM: Honesty builds trust, reduces walk-outs', 'Wait time tracking', ''],
    ['Provider-to-front-desk handoff', 'Provider walks patient to front desk side-by-side, names next service', 'Rebooking Rate → Unique Patients → Revenue', 'VERY HIGH: Provider endorsement of next visit 3x more effective than front desk alone', 'Handoff compliance rate', ''],
    ['Rebook at checkout', '"Single most important retention activity" — state ideal return date as recommendation', 'Rebooking Rate → Unique Patients → Revenue', 'VERY HIGH: Target >65% rebooking rate at checkout', 'Rebooking rate at checkout', ''],
    ['Check loyalty rewards before payment', 'Review Allē wallet, points, coupons before processing payment', 'Patient Satisfaction → Retention, Collections %', 'MEDIUM: Prevents negative experience of missed rewards', 'Rewards application rate', ''],
    ['Express warmth at departure', 'Genuine compliment about treatment decision, encourage future contact', 'Patient Satisfaction → Reviews → New Patients', 'LOW-MEDIUM: Last impression as important as first', 'Patient NPS', ''],
]

for i, row_data in enumerate(checkin_data):
    for j, val in enumerate(row_data):
        ws.cell(row=r, column=j+1, value=val)
    drow(ws, r, len(headers), alt=(i%2==0))
    ws.cell(row=r, column=1).font = header_font
    ws.cell(row=r, column=6).fill = yellow_fill
    ws.cell(row=r, column=6).font = fill_font
    r += 1

r += 2
# Section C: Scheduling
ws.cell(row=r, column=1, value='C. SCHEDULING → UTILIZATION & REVENUE').font = subtitle_font
r += 1
for j, h in enumerate(headers):
    ws.cell(row=r, column=j+1, value=h)
hdr(ws, r, len(headers))
r += 1

sched_data = [
    ['Offer first available proactively', '"We have Monday 2pm or Tuesday 9am — which works better?"', 'Utilization Rate → Rev Per Provider Hour', 'MEDIUM: Reduces time-to-appointment, fills gaps', 'Days-to-next-appt average', ''],
    ['Target <48hr for non-surgical', 'Non-surgical consultations within 48 hours of inquiry', 'Unique Patients → Revenue', 'HIGH: Every day of delay reduces conversion', 'Time-to-appointment tracking', ''],
    ['Explain confirmation process at booking', 'Two confirmations: 48hr + 24hr. Confirm text permission.', 'Cancel Rate → Utilization → Revenue', 'HIGH: Clear expectations reduce day-of cancellations', 'Confirmation completion rate', ''],
    ['Pre-arrival registration', 'Forms completed before arrival; arrive 15 min early if not', 'Check-in Speed → Provider Start Time → Utilization', 'MEDIUM: Reduces wasted provider time at start of appt', 'Pre-registration completion rate', ''],
    ['Schedule 2-3 weeks ahead at checkout', 'Book follow-up before patient leaves. State specific date.', 'Rebooking Rate → Utilization → Revenue Predictability', 'VERY HIGH: Pre-booked patients show up 3x more reliably', 'Forward booking rate', ''],
    ['Monitor shift edge gaps', 'Flag any provider-day where first appt starts 2+ hrs after shift', 'Utilization Rate → Rev Per Provider Hour', 'HIGH: 3+ hr morning gap HALVES rev/hr from $269 to $138', 'Morning/evening gap report', ''],
    ['Reduce wasteful shifts', 'Thu has 27.4% waste rate. Consider half-day shifts on slow days.', 'Net Provider Hours → Rev Per Scheduled Hour', 'VERY HIGH: 26,845 wasted hrs in 4 wks = ~$3.1M unrealized', 'Shifts with <2 booked hours', ''],
]

for i, row_data in enumerate(sched_data):
    for j, val in enumerate(row_data):
        ws.cell(row=r, column=j+1, value=val)
    drow(ws, r, len(headers), alt=(i%2==0))
    ws.cell(row=r, column=1).font = header_font
    ws.cell(row=r, column=6).fill = yellow_fill
    ws.cell(row=r, column=6).font = fill_font
    r += 1

widths = [28, 50, 30, 35, 28, 25]
for i, w in enumerate(widths):
    ws.column_dimensions[get_column_letter(i+1)].width = w

# ═══════════════════════════════════════════════════
# SHEET: Allē Revenue Engine
# ═══════════════════════════════════════════════════
if 'Alle Revenue Engine' in wb.sheetnames:
    del wb['Alle Revenue Engine']
ws2 = wb.create_sheet('Alle Revenue Engine')
ws2.sheet_properties.tabColor = '4A7C6F'

r = 1
ws2.cell(row=r, column=1, value='Allē Management → KPI Impact Map').font = title_font
r += 1
ws2.cell(row=r, column=1, value='How Allē program management directly drives injectable revenue, patient retention, and collections efficiency.').font = small_font
r += 2

# Connection diagram
ws2.cell(row=r, column=1, value='ALLĒ → KPI CAUSAL CHAIN').font = subtitle_font
r += 1
chain_headers = ['Allē Activity', 'Cadence', 'Direct KPI Impact', 'Secondary KPI Impact', 'Revenue Mechanism']
for j, h in enumerate(chain_headers):
    ws2.cell(row=r, column=j+1, value=h)
hdr(ws2, r, len(chain_headers))
r += 1

chains = [
    ['100% Sign-Up Rate', 'Daily (every patient)', 'Allē Enrollment → Flash Scan Eligibility', 'Injectable Revenue, Dual-Treatment Rebates', 'Enrolled patients earn points → return for redemption → recurring revenue cycle'],
    ['Flash Scanning at Checkout', 'Daily (every injectable)', 'Points Accumulation → Return Visit Motivation', 'Rebooking Rate, Patient Lifetime Value', 'Scanned patients accumulate value. Unscanned patients lose motivation to return.'],
    ['Coupon Wallet Check', 'Daily (every checkout)', 'Collections %, Patient Satisfaction', 'Retention Rate, NPS', 'Applying available rewards prevents negative checkout experience and increases perceived value'],
    ['Patient 360 Report Review', 'Weekly', 'Lapsed Patient Reactivation → Unique Patients', 'Revenue, Utilization, Injectable Revenue', 'Filling provider gaps with lapsed patients = immediate revenue recovery at zero acquisition cost'],
    ['Expiring Points Outreach', 'Weekly', 'Rebooking Rate, Unique Patients', 'Revenue, Injectable %, Utilization', 'Expiring points create urgency. Warm leads with built-in motivation to return.'],
    ['Dual-Treatment Conversion', 'Monthly target', 'Injectable Revenue, Filler %, NTX Revenue', 'Avg Revenue Per Patient, Rebate Tier', 'Neuro + Filler in same month → higher Allergan rebates → lower COGS → better margins'],
    ['Cherry Payment Plans', 'Monthly target', 'Collections %, Treatment Acceptance Rate', 'Revenue Per Patient, Injectable Revenue', 'Financing removes price objection → larger treatment plans → immediate cash collection'],
    ['Lapsed Client Re-Engagement', 'Monthly campaign', 'Unique Patients, Utilization', 'Revenue, Injectable Revenue', 'Reactivation costs 5-10x less than acquisition. Monthly outreach = highest ROI activity.'],
    ['Rebate Tier Optimization', 'Monthly review', 'Allergan Rebates → Margin', 'Profitability, COGS Reduction', 'Dual-treatment + Cherry volume → tier advancement → lower per-unit cost on all Allergan products'],
]

for i, row_data in enumerate(chains):
    for j, val in enumerate(row_data):
        ws2.cell(row=r, column=j+1, value=val)
    drow(ws2, r, len(chain_headers), alt=(i%2==0))
    ws2.cell(row=r, column=1).font = header_font
    r += 1

r += 2
# Allē KPI Targets
ws2.cell(row=r, column=1, value='ALLĒ PERFORMANCE BENCHMARKS').font = subtitle_font
r += 1
kpi_headers = ['Allē Metric', 'Target', 'Critical Threshold', 'Measurement Method', 'Frequency', 'YOUR: Current Status']
for j, h in enumerate(kpi_headers):
    ws2.cell(row=r, column=j+1, value=h)
hdr(ws2, r, len(kpi_headers))
r += 1

alle_kpis = [
    ['Allē Sign-Up Rate (injectable patients)', '100%', '<80% = retrain front desk', 'Enrolled patients / injectable patients', 'Daily', ''],
    ['Flash Scan Rate', '100%', '<90% = retrain checkout workflow', 'Scans / eligible treatments', 'Weekly audit', ''],
    ['Coupon Wallet Check Rate', '100%', '<90% = process gap', 'Observed compliance at checkout', 'Weekly spot-check', ''],
    ['Patient 360 Report Reviewed', 'Weekly', 'Missed 2+ weeks = critical', 'PM confirms review in weekly standup', 'Weekly', ''],
    ['Lapsed Patient Outreach (6+ months)', '100% of list contacted monthly', '<50% = revenue leakage', 'Contacts made / lapsed patients identified', 'Monthly', ''],
    ['Dual-Treatment Conversion Rate', 'Practice-specific target', '<10% = missed rebate opportunity', 'Patients with both NTX + filler in calendar month', 'Monthly', ''],
    ['Cherry Application Rate', 'Growing MoM', 'Flat or declining = training gap', 'New applications / consultations', 'Monthly', ''],
    ['Expiring Points Recovery Rate', '>60% redeemed before expiry', '<40% = outreach failure', 'Points redeemed / points expiring', 'Weekly', ''],
    ['Promotional Pull-Through', '>50% redemption', '<30% = campaign disconnect', 'Coupons redeemed / coupons issued', 'Monthly', ''],
]

for i, row_data in enumerate(alle_kpis):
    for j, val in enumerate(row_data):
        ws2.cell(row=r, column=j+1, value=val)
    drow(ws2, r, len(kpi_headers), alt=(i%2==0))
    ws2.cell(row=r, column=1).font = header_font
    ws2.cell(row=r, column=6).fill = yellow_fill
    ws2.cell(row=r, column=6).font = fill_font
    r += 1

r += 2
ws2.cell(row=r, column=1, value='KEY INSIGHT: 50% of first-time injectable clients who do not return within 6 months never become repeat patients. Proactive Allē management is the #1 retention lever.').font = Font(name='Arial', italic=True, color=RED, size=10)
ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)

widths2 = [35, 25, 28, 35, 15, 25]
for i, w in enumerate(widths2):
    ws2.column_dimensions[get_column_letter(i+1)].width = w

# ═══════════════════════════════════════════════════
# SHEET: Provider Productivity Deep Dive
# ═══════════════════════════════════════════════════
if 'Provider Productivity' in wb.sheetnames:
    del wb['Provider Productivity']
ws3 = wb.create_sheet('Provider Productivity')
ws3.sheet_properties.tabColor = RED

r = 1
ws3.cell(row=r, column=1, value='Provider Productivity → Revenue & Cost Impact').font = title_font
r += 1
ws3.cell(row=r, column=1, value='How individual provider metrics cascade to location-level revenue, labor efficiency, and profitability.').font = small_font
r += 2

# Correlation findings
ws3.cell(row=r, column=1, value='A. CORRELATION ANALYSIS FINDINGS').font = subtitle_font
r += 1
corr_headers = ['Provider Metric', 'Correlation With', 'Strength', 'Direction', 'Interpretation']
for j, h in enumerate(corr_headers):
    ws3.cell(row=r, column=j+1, value=h)
hdr(ws3, r, len(corr_headers))
r += 1

correlations = [
    ['Injectable % of Revenue', 'Revenue Per Patient', 'Strong (r=0.72)', 'Positive', 'Providers with higher injectable share generate ~2x more revenue per patient than non-injectable-focused providers'],
    ['Injectable % of Revenue', 'Revenue Per Billable Hour', 'Strong (r=0.68)', 'Positive', 'Injectable procedures generate $1,331-2,237/hr vs $605-663/hr for non-injectable services'],
    ['Botox Units Per Appt', 'Neurotoxin Revenue', 'Very Strong (r=0.85)', 'Positive', 'Each additional 10 units ≈ $120-150 more revenue per appointment. Under-dosing (<40 units) is a direct revenue leak.'],
    ['Filler Syringes Per Appt', 'Filler Revenue', 'Very Strong (r=0.88)', 'Positive', 'Each additional syringe ≈ $600-800 more revenue. NTX+Filler combo generates 3.5x the rev/hr of laser.'],
    ['Utilization Rate', 'Revenue Per Scheduled Hour', 'Very Strong (r=0.82)', 'Positive', '<30% util = $98/billable hr. 70-85% = $482/billable hr. Sweet spot: 70-85%.'],
    ['Blockout Hours', 'Revenue Per Scheduled Hour', 'Strong (r=-0.65)', 'Negative', '>10 hrs blockout/wk drops efficiency to $125/scheduled hr — almost certainly unprofitable'],
    ['Morning Gap (hrs before first appt)', 'Utilization Rate', 'Strong (r=-0.58)', 'Negative', '3+ hr morning gap HALVES utilization (85.4% → 40.6%) and rev/hr ($269 → $138)'],
    ['Cancel Rate', 'Unique Patients', 'Moderate (r=-0.45)', 'Negative', 'Each 1% increase in cancel rate ≈ 2-3 fewer patients per week per location'],
    ['No-Show Rate', 'Revenue (unrecoverable)', 'Moderate (r=-0.40)', 'Negative', 'No-shows are worse than cancellations — no opportunity to backfill the slot'],
    ['Dual Treatment (NTX+Filler)', 'Revenue Per Patient', 'Very Strong (r=0.78)', 'Positive', 'Patients receiving both NTX + filler generate $1,528 vs $555 (NTX only) or $1,088 (filler only)'],
    ['Patient Volume (unique patients)', 'Revenue', 'Very Strong (r=0.91)', 'Positive', 'Volume is the #1 driver of total revenue. But volume without yield = low rev/patient.'],
    ['Retail %', 'Revenue Per Patient', 'Weak (r=0.15)', 'Positive', 'Retail adds margin but doesn\'t meaningfully move rev/patient. More important for margin than topline.'],
]

for i, row_data in enumerate(correlations):
    for j, val in enumerate(row_data):
        ws3.cell(row=r, column=j+1, value=val)
    drow(ws3, r, len(corr_headers), alt=(i%2==0))
    ws3.cell(row=r, column=1).font = header_font
    r += 1

r += 2
# Provider Action Triggers
ws3.cell(row=r, column=1, value='B. PROVIDER-LEVEL ACTION TRIGGERS').font = subtitle_font
r += 1
ws3.cell(row=r, column=1, value='When a provider\'s metrics cross these thresholds, the skill triggers specific recommendations.').font = small_font
r += 1

trig_headers = ['Trigger Condition', 'Severity', 'Recommended Action', 'Who Delivers', 'Timeline', 'YOUR: Override/Notes']
for j, h in enumerate(trig_headers):
    ws3.cell(row=r, column=j+1, value=h)
hdr(ws3, r, len(trig_headers))
r += 1

triggers = [
    ['Botox units < 30 (3+ weeks)', 'CRITICAL', 'Immediate 1:1 with Medical Director. Review full-face assessment protocol. Shadow top performer.', 'Medical Director', '1 week', ''],
    ['Botox units 30-40 (declining trend)', 'IMPORTANT', 'Group training on dosing protocols. Review before/after portfolios. Set 40-unit floor target.', 'Medical Director', '2 weeks', ''],
    ['Filler syringes/appt < 1.0', 'IMPORTANT', 'Treatment plan depth training. Review consultation videos. Practice multi-area assessments.', 'Medical Director', '2-3 weeks', ''],
    ['Injectable % < 20% (for injector role)', 'CRITICAL', 'Review appointment types being booked. Are injectable slots being filled with non-injectable services?', 'PM + Scheduling Lead', '1 week', ''],
    ['Revenue/billable hour < $200', 'CRITICAL', 'Combined issue: low yield AND/OR low utilization. Check service mix first, then schedule density.', 'PM', '2 weeks', ''],
    ['Utilization < 50% (3+ weeks)', 'CRITICAL', 'Reduce scheduled hours or shift to half-days. Fill gaps with lapsed patient recalls.', 'PM', '1 week', ''],
    ['Morning gap > 2 hours (recurring)', 'IMPORTANT', 'Move shift start to match first appointment. Wasted morning hours are pure labor cost.', 'PM', 'Immediate', ''],
    ['Evening gap > 2 hours (recurring)', 'OPTIMIZE', 'Shift end time earlier or run last-minute booking campaign for end-of-day slots.', 'PM + Marketing', '1-2 weeks', ''],
    ['Cancel rate > 10% (provider-specific)', 'IMPORTANT', 'Review which patients are cancelling. Provider communication issue or scheduling conflict?', 'PM', '2 weeks', ''],
    ['Collections % < 50% (provider-specific)', 'IMPORTANT', 'Review payment types. High package redemption? Discuss point-of-sale collection training.', 'PM + Front Desk Lead', '2 weeks', ''],
    ['No injectable revenue (injector role, 2+ weeks)', 'CRITICAL', 'Appointment type audit. Why is this injector not seeing injectable patients? Schedule template issue?', 'PM', 'Immediate', ''],
    ['Revenue/patient declining 3+ weeks', 'IMPORTANT', 'Review treatment plans. Is provider under-treating? Shorter consultations? Missing cross-sell?', 'Medical Director', '2 weeks', ''],
]

for i, row_data in enumerate(triggers):
    for j, val in enumerate(row_data):
        ws3.cell(row=r, column=j+1, value=val)
    drow(ws3, r, len(trig_headers), alt=(i%2==0))
    ws3.cell(row=r, column=1).font = header_font
    ws3.cell(row=r, column=6).fill = yellow_fill
    ws3.cell(row=r, column=6).font = fill_font
    # Color severity
    sev = row_data[1]
    sev_cell = ws3.cell(row=r, column=2)
    if sev == 'CRITICAL': sev_cell.fill = PatternFill('solid', fgColor=RED_BG); sev_cell.font = Font(name='Arial', bold=True, color=RED, size=10)
    elif sev == 'IMPORTANT': sev_cell.fill = PatternFill('solid', fgColor=YELLOW); sev_cell.font = Font(name='Arial', bold=True, color=GOLD, size=10)
    else: sev_cell.fill = PatternFill('solid', fgColor=BLUE_BG); sev_cell.font = Font(name='Arial', bold=True, color='5B8CB9', size=10)
    r += 1

r += 2
# Revenue Impact Model
ws3.cell(row=r, column=1, value='C. SERVICE MIX → REVENUE PER HOUR (from data analysis)').font = subtitle_font
r += 1
svc_headers = ['Service Type', 'Avg Revenue', 'Avg Duration', 'Rev/Hour', '% of Total Rev', 'Skill Implication']
for j, h in enumerate(svc_headers):
    ws3.cell(row=r, column=j+1, value=h)
hdr(ws3, r, len(svc_headers))
r += 1

svc_data = [
    ['NTX + Filler Combo', '$1,528', '41 min', '$2,237/hr', '14.5%', 'HIGHEST VALUE. Every combo appt replaces 3.5 laser hours of revenue. Prioritize dual-treatment conversion.'],
    ['Other Injectable', '$998', '36 min', '$1,663/hr', '5.4%', 'High value. Includes Kybella, biostimulators. Encourage provider training on newer modalities.'],
    ['Filler Only', '$1,088', '47 min', '$1,389/hr', '7.9%', 'Strong value but could be combo. Cross-sell NTX at every filler appointment.'],
    ['Neurotoxin Only', '$555', '25 min', '$1,331/hr', '28.9%', 'Best volume driver. Quick turns. Ensure units ≥40 per appointment to maximize.'],
    ['Skin Rejuvenation', '$830', '48 min', '$1,038/hr', '10.1%', 'Good supplementary service. Less yield than injectables but fills schedule.'],
    ['Laser Hair Reduction', '$309', '28 min', '$663/hr', '18.3%', 'High volume, low yield. 8 hrs of laser = 3 hrs of combo injectable revenue.'],
    ['Body Contouring', '$932', '88 min', '$636/hr', '2.4%', 'Longest duration, lowest rev/hr. Schedule in gaps, not in prime injectable time.'],
    ['Other Services', '$484', '48 min', '$605/hr', '12.5%', 'Catch-all. Review what\'s in here — may include low-value services consuming prime time.'],
]

for i, row_data in enumerate(svc_data):
    for j, val in enumerate(row_data):
        ws3.cell(row=r, column=j+1, value=val)
    drow(ws3, r, len(svc_headers), alt=(i%2==0))
    ws3.cell(row=r, column=1).font = header_font
    r += 1

ws3.cell(row=r+1, column=1, value='KEY: Converting 1 laser hour to 1 injectable hour = $668+ additional revenue. The skill should flag providers with high laser/low injectable mix.').font = red_font
ws3.merge_cells(start_row=r+1, start_column=1, end_row=r+1, end_column=6)

widths3 = [32, 14, 35, 18, 15, 45]
for i, w in enumerate(widths3):
    ws3.column_dimensions[get_column_letter(i+1)].width = w

# ═══════════════════════════════════════════════════
# SHEET: PM Skill Architecture
# ═══════════════════════════════════════════════════
if 'PM Skill Architecture' in wb.sheetnames:
    del wb['PM Skill Architecture']
ws4 = wb.create_sheet('PM Skill Architecture')
ws4.sheet_properties.tabColor = NAVY

r = 1
ws4.cell(row=r, column=1, value='PM Skill — 5-Layer Analysis Architecture').font = title_font
r += 1
ws4.cell(row=r, column=1, value='How the daily automated analysis processes data through 5 layers to produce actionable recommendations per location and provider.').font = small_font
r += 2

arch_headers = ['Layer', 'Function', 'Data Sources', 'Output', 'Example']
for j, h in enumerate(arch_headers):
    ws4.cell(row=r, column=j+1, value=h)
hdr(ws4, r, len(arch_headers))
r += 1

layers = [
    ['1. Trend Detection', 'Identify which KPIs are rising/falling vs prior 4 weeks and vs peers', 'weekly-metrics, weekly-ops, weekly-btx, weekly-utilization, weekly-provider-hours', '"Revenue down 12% WoW. 3rd consecutive decline."', 'Zona Rosa revenue: $26.6K → $21.2K → $12.9K (3-week slide)'],
    ['2. Causal Diagnosis', 'Trace WHY a KPI changed using the causal map', 'KPI Causal Map + correlation coefficients + provider data', '"Revenue decline caused by: 23% drop in unique patients + 15% drop in avg rev/patient"', 'Unique patients dropped because cancel rate spiked to 8.3%'],
    ['3. Provider Spotlight', 'Identify which providers are dragging down or lifting each KPI', 'weekly-inj-rev-provider, weekly-btx-provider, weekly-syringe-provider, weekly-metrics-provider', '"Jamie Taylor: btx units dropped from 44→29 over 3 weeks (below 40-unit floor)"', 'Raina Murray carrying 40% of injectable revenue; Jamie declining'],
    ['4. Action Prescription', 'Recommend specific actions from playbooks + operational levers', 'Front Office Playbook + Allē SOP + Operational Levers + Practice Playbooks', '"IMMEDIATE: Schedule 1:1 with Jamie Taylor on full-face assessment. Review dual-treatment opportunity."', 'Reference specific Allē/Front Office protocols with expected impact'],
    ['5. Schedule Optimization', 'Flag wasteful shifts, morning gaps, over/under-booked providers', 'weekly-provider-hours + utilization + appt patterns', '"Thursday 3 providers scheduled, only 8 appointments. Recommend half-day for 1 provider."', 'Shift Provider X start from 8am to 10am (first appt at 10:15am)'],
]

for i, row_data in enumerate(layers):
    for j, val in enumerate(row_data):
        ws4.cell(row=r, column=j+1, value=val)
    drow(ws4, r, len(arch_headers), alt=(i%2==0))
    ws4.cell(row=r, column=1).font = Font(name='Arial', bold=True, color=WHITE, size=11)
    ws4.cell(row=r, column=1).fill = navy_fill
    r += 1

r += 2
ws4.cell(row=r, column=1, value='DAILY OUTPUT FORMAT').font = subtitle_font
r += 1
format_data = [
    ['Section', 'Content'],
    ['Header', 'Location name, date, period analyzed, overall health score (1-10)'],
    ['🔴 Critical Alerts', 'KPIs below critical thresholds. Requires immediate action this week.'],
    ['📊 Trend Summary', '4-week trend for each KPI with directional arrows and peer comparison'],
    ['👤 Provider Spotlight', 'Top 3 improving + top 3 declining providers with specific metrics'],
    ['💡 Recommendations', '3-5 prioritized actions with specific names, protocols, and expected impact'],
    ['📅 This Week Predictions', 'Projected revenue based on bookings, historical patterns, and current trends'],
    ['📅 Next Week Outlook', 'Forward-looking based on schedule density, seasonal patterns, and trend direction'],
]

for i, row_data in enumerate(format_data):
    for j, val in enumerate(row_data):
        ws4.cell(row=r, column=j+1, value=val)
    if i == 0:
        hdr(ws4, r, 2)
    else:
        drow(ws4, r, 2, alt=(i%2==0))
        ws4.cell(row=r, column=1).font = header_font
    r += 1

widths4 = [25, 30, 45, 45, 50]
for i, w in enumerate(widths4):
    ws4.column_dimensions[get_column_letter(i+1)].width = w

# Save
wb.save(outfile)
print('Saved to:', outfile)
print('Final sheets:', wb.sheetnames)
