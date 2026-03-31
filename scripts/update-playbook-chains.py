import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY = '041E42'
GOLD = 'B9975B'
CREAM = 'FAF8F7'
WHITE = 'FFFFFF'
TAUPE = 'E4D5D3'
YELLOW = 'FFF3CD'
RED = 'C0392B'
GREEN = '1A6B3C'
RED_BG = 'FFE0E0'
GREEN_BG = 'E0FFE0'
BLUE_BG = 'E0E8FF'

navy_fill = PatternFill('solid', fgColor=NAVY)
cream_fill = PatternFill('solid', fgColor=CREAM)
white_fill = PatternFill('solid', fgColor=WHITE)
yellow_fill = PatternFill('solid', fgColor=YELLOW)
green_bg = PatternFill('solid', fgColor=GREEN_BG)

navy_font = Font(name='Arial', bold=True, color=WHITE, size=11)
header_font = Font(name='Arial', bold=True, color=NAVY, size=10)
body_font = Font(name='Arial', color='2A1F28', size=10)
small_font = Font(name='Arial', color='948794', size=9)
fill_font = Font(name='Arial', color='0000FF', size=10)
title_font = Font(name='Arial', bold=True, color=NAVY, size=14)
subtitle_font = Font(name='Arial', bold=True, color=GOLD, size=12)
bold_navy = Font(name='Arial', bold=True, color=NAVY, size=10)
red_italic = Font(name='Arial', italic=True, color=RED, size=9)
green_italic = Font(name='Arial', italic=True, color=GREEN, size=9)

thin_border = Border(
    left=Side(style='thin', color=TAUPE), right=Side(style='thin', color=TAUPE),
    top=Side(style='thin', color=TAUPE), bottom=Side(style='thin', color=TAUPE),
)

def hdr(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = navy_fill; cell.font = navy_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

def drow(ws, row, cols, alt=False):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = cream_fill if alt else white_fill
        cell.font = body_font; cell.border = thin_border
        cell.alignment = Alignment(vertical='top', wrap_text=True)

wb = openpyxl.load_workbook('AMP_KPI_Operational_Playbook_v3.xlsx')

# ═══════════════════════════════════════════════════
# SHEET: Metric Chains (THE core sheet)
# ═══════════════════════════════════════════════════
if 'Metric Chains' in wb.sheetnames:
    del wb['Metric Chains']
ws = wb.create_sheet('Metric Chains', 0)  # Insert as FIRST sheet
ws.sheet_properties.tabColor = RED

r = 1
ws.cell(row=r, column=1, value='How Each Metric Mechanically Moves Another Metric').font = title_font
r += 1
ws.cell(row=r, column=1, value='Quantified from 2,962 location-weeks of AMP data. These are the exact levers the PM Skill uses to diagnose problems and prescribe actions.').font = small_font
r += 2

# ═══════ SECTION 1: The Revenue Equation ═══════
ws.cell(row=r, column=1, value='1. THE REVENUE EQUATION').font = subtitle_font
r += 1
ws.cell(row=r, column=1, value='Revenue = Unique Patients × Avg Revenue Per Patient').font = Font(name='Arial', bold=True, color=NAVY, size=12)
r += 1
ws.cell(row=r, column=1, value='This is the fundamental equation. Every other metric feeds into one of these two components.').font = small_font
r += 2

headers = ['Input Metric', 'Quantified Impact', 'Mechanism', 'Tasks to Move This Metric', 'Expected $ Impact Per Location']
for j, h in enumerate(headers):
    ws.cell(row=r, column=j+1, value=h)
hdr(ws, r, len(headers))
r += 1

rev_chains = [
    ['+1 unique patient', '+$453 weekly revenue', 'Direct: more patients = more revenue. This is the #1 revenue driver (r=0.80).', '1. Reduce cancel rate (each 1% drop = +2 patients)\n2. Reduce no-show rate (each 1% drop = +8.5 patients)\n3. Increase rebooking at checkout (target >65%)\n4. Allē lapsed patient outreach (6+ month no-visit)\n5. Pre-consultation calls (increases show rate)\n6. Speed of response <1 min (+391% conversion)', 'Moving from bottom quartile (39 patients) to top quartile (189 patients) = +$68,778/wk'],
    ['+$1 avg rev per patient', '+$56 weekly revenue', 'Multiplier: same patients generate more per visit. Second-largest driver (r=0.52).', '1. Increase injectable % (higher-yield services)\n2. Full-face botox assessments (push units >40)\n3. Treatment plan depth (multi-area filler)\n4. Cross-sell NTX + Filler in same visit (+$1,528 avg)\n5. CIQ at check-in (expands consultation scope)\n6. Provider endorsement of comprehensive plans', 'Moving from bottom quartile ($274/pt) to top quartile ($743/pt) = +$49,315/wk'],
]

for i, row_data in enumerate(rev_chains):
    for j, val in enumerate(row_data):
        ws.cell(row=r, column=j+1, value=val)
    drow(ws, r, len(headers), alt=(i%2==0))
    ws.cell(row=r, column=1).font = bold_navy
    r += 1

r += 2
# ═══════ SECTION 2: Patient Volume Chain ═══════
ws.cell(row=r, column=1, value='2. PATIENT VOLUME CHAIN (What Drives Unique Patients)').font = subtitle_font
r += 1

vol_chains = [
    ['-1% cancel rate', '+2.0 patients/week', 'Cancelled appointments = lost volume. AMP avg cancel rate: 5.1%. Each 1% reduction recovers 2 patients.', '1. 48hr + 24hr confirmation texts (cheapest, fastest)\n2. Credit card hold policy for high-value appointments\n3. Waitlist management to backfill cancellations\n4. Proactive delay communication (reduces walk-outs)\n5. Explain confirmation process clearly at booking\n6. Same-day callback for cancellation reschedule', 'Reducing cancel from 9% (bottom) to 2% (top) = +$7,296/wk'],
    ['-1% no-show rate', '+8.6 patients/week', 'No-shows are 4x worse than cancellations per % point. AMP avg: 2.7%. No-shows cannot be backfilled.', '1. SMS reminders at 72hr, 24hr, 2hr before appointment\n2. No-show fee policy ($50-100)\n3. Personal phone call for high-value appointments\n4. Pre-consultation call (creates commitment)\n5. Deposit requirement for new patients\n6. Allē Flash reminder (gives patient reason to show)', 'Each 1% reduction = +8.6 patients = +$3,897/wk revenue'],
    ['+10% rebooking rate', '+6-8 patients/week (recurring)', 'Rebooking at checkout = guaranteed future volume. Compounding: rebookers return repeatedly.', '1. Provider walks patient to front desk (3x more effective)\n2. State specific return date as recommendation\n3. "Your next treatment should be around [Month/Date]"\n4. Check Allē wallet for expiring points (creates urgency)\n5. Schedule 2-3 weeks ahead before patient leaves\n6. Follow-up call within 48hrs if not rebooked', 'Target >65% rebooking rate. Moving from 40% to 65% = ~$3,000-5,000/wk'],
    ['+1 new patient inquiry converted', '+$453 first-visit revenue', 'New patient conversion is the top-of-funnel driver. AMP digital response speed matters enormously.', '1. Answer phone before 3 rings\n2. Digital response <1 minute (+391% conversion vs slow)\n3. Credential the 3 P\'s (Practice, Provider, Procedure)\n4. Pre-consultation call within 48hrs of booking\n5. Always offer next step (never end call without action)\n6. Capture contact info on chat/digital immediately', 'Each converted inquiry = $453 immediate + lifetime value'],
]

for j, h in enumerate(headers):
    ws.cell(row=r, column=j+1, value=h)
hdr(ws, r, len(headers))
r += 1

for i, row_data in enumerate(vol_chains):
    for j, val in enumerate(row_data):
        ws.cell(row=r, column=j+1, value=val)
    drow(ws, r, len(headers), alt=(i%2==0))
    ws.cell(row=r, column=1).font = bold_navy
    r += 1

r += 2
# ═══════ SECTION 3: Revenue Per Patient Chain ═══════
ws.cell(row=r, column=1, value='3. REVENUE PER PATIENT CHAIN (What Drives Yield)').font = subtitle_font
r += 1

yield_chains = [
    ['+1% injectable share of revenue', '+$0.89 rev per patient', 'Injectable services generate 2-3.5x more rev/hr than non-injectable. Shifting mix toward injectables lifts yield.', '1. Provider training on injectable assessment\n2. CIQ at check-in (captures injectable opportunity)\n3. Cross-sell NTX at every filler visit and vice versa\n4. Allē dual-treatment promotion\n5. Schedule injectable patients in prime time slots\n6. Marketing focus on injectable services', 'Moving injectable % from 32% (bottom) to 85% (top) = +$9,248/wk'],
    ['+10 botox units per appointment', '+$3,976 NTX rev per location-week', 'Under-dosing (below 40 units) is the most common revenue leak. Proper dosing = better results AND more revenue.', '1. Full-face assessment training (not just treatment area)\n2. Dosing protocol standardization (40-unit floor)\n3. Medical Director 1:1 coaching for providers <40 units\n4. Before/after portfolio showing full-face results\n5. Patient education on proper dosing for longevity\n6. Shadow top-performing injector for one session', 'Moving from 36 units (bottom) to 69 units (top) = +$29,822/wk'],
    ['+1 filler syringe per appointment', '+$2,871 filler rev per location-week', 'Multi-syringe treatments generate dramatically more per visit. NTX+Filler combo = $2,237/hr (3.5x laser rate).', '1. Treatment plan depth training (multi-area approach)\n2. Before/after portfolio showing multi-syringe results\n3. Financing promotion (Cherry/PatientFi) — removes price barrier\n4. Provider confidence building — demonstrate outcomes\n5. Consultation scripts for comprehensive assessment\n6. Allē dual-treatment rebate positioning', 'Each additional syringe per filler appt adds ~$600-800 revenue'],
    ['+1% dual treatment rate (NTX+Filler same visit)', '+$1.37 rev per patient', 'Patients receiving both NTX + filler generate $1,528 avg vs $555 (NTX only). Massive yield multiplier.', '1. Ask "Have you considered [NTX/filler] as well?" at every injectable consult\n2. Package pricing for combo treatments\n3. Allē dual-treatment rebate education for patients\n4. Schedule combo appointments with adequate time\n5. Provider training on combo assessment\n6. Cherry financing for higher-ticket combo plans', 'Moving dual treatment from 10% to 30% = significant rev/patient lift'],
    ['+1% retail share', '+$1.65 rev per patient', 'Retail is low-effort, high-margin. Small impact on rev/patient but meaningful margin contribution.', '1. Provider post-treatment product recommendation (specific product, not generic)\n2. Product bundles with services at checkout\n3. Product displays at checkout counter\n4. Staff training on top 5 products per treatment type\n5. Check-out script: "Dr. X recommends [Product] for your results"\n6. Sample program to drive trial', 'Retail goal: 7.5% of revenue. Moving from 1.4% to 16% = +$3,507/wk'],
]

for j, h in enumerate(headers):
    ws.cell(row=r, column=j+1, value=h)
hdr(ws, r, len(headers))
r += 1

for i, row_data in enumerate(yield_chains):
    for j, val in enumerate(row_data):
        ws.cell(row=r, column=j+1, value=val)
    drow(ws, r, len(headers), alt=(i%2==0))
    ws.cell(row=r, column=1).font = bold_navy
    r += 1

r += 2
# ═══════ SECTION 4: Efficiency Chain ═══════
ws.cell(row=r, column=1, value='4. EFFICIENCY CHAIN (Revenue & Collections Per Provider Hour)').font = subtitle_font
r += 1
ws.cell(row=r, column=1, value='Rev/Billable Hour = Revenue ÷ (Scheduled Hours - Blockout Hours). This is the cost containment + profitability metric.').font = small_font
r += 1

eff_chains = [
    ['+1% utilization rate', '+$0.60 rev per billable hour', 'Utilization = booked time / available time. More bookings in same hours = more revenue per hour paid.', '1. Optimize scheduling templates (minimize gaps)\n2. Reduce morning gaps (3+hr gap halves utilization)\n3. Reduce evening gaps (shift end earlier if last appt >2hr early)\n4. Last-minute booking promos for open slots\n5. Allē lapsed patient recalls for empty slots\n6. Adjust provider shifts to match demand patterns', 'Moving utilization from 65% (bottom) to 94% (top) = +$21,000/wk revenue'],
    ['-1 blockout hour per week', '+$2.29 rev per scheduled hour', 'Blockout = paid time that can\'t generate revenue. >10 hrs/wk drops to $125/scheduled hr (unprofitable).', '1. Audit all blockout types (admin, lunch, training, meetings)\n2. Move admin tasks to before/after patient hours\n3. Shorten lunch blocks to 30 min where possible\n4. Schedule provider training outside peak booking hours\n5. Convert recurring blockouts to bookable time\n6. Flag any provider with >5 hrs/wk blockout for review', 'Reducing blockout from 10+ hrs to <2 hrs = $125→$437/scheduled hr'],
    ['Eliminate 1 wasteful shift per week', '+$335 rev per eliminated shift', 'A "wasteful shift" = <2 booked hours. Thursday worst at 27.4% waste. Pure labor cost with minimal revenue.', '1. Review shifts with <2 booked hours over past 4 weeks\n2. Convert wasteful full-day shifts to half-days\n3. Eliminate Thursday shifts for providers with chronic low volume\n4. Use waitlist to fill shifts before they become wasteful\n5. Shift provider hours from low-demand to high-demand days\n6. Consider shared/float providers across locations', 'Eliminating 26,845 wasted hrs = ~$3.1M unrealized revenue in 4 wks'],
    ['Close morning gap by 1 hour', '+$131 rev per billable hour', '3+ hr morning gap = 40.6% utilization, $138/billable hr. No gap = 85.4%, $269/hr. Gap is pure waste.', '1. Match shift start to first appointment (not arbitrary 8am)\n2. If first appt is 10am, start shift at 9:45am\n3. Backfill morning gaps with admin tasks only if unfillable\n4. Run morning-specific promotions for early slots\n5. Schedule follow-up/quick appointments at shift start\n6. Track gap report weekly per provider', 'Closing a 3-hr morning gap = nearly doubles productivity for that shift'],
]

for j, h in enumerate(headers):
    ws.cell(row=r, column=j+1, value=h)
hdr(ws, r, len(headers))
r += 1

for i, row_data in enumerate(eff_chains):
    for j, val in enumerate(row_data):
        ws.cell(row=r, column=j+1, value=val)
    drow(ws, r, len(headers), alt=(i%2==0))
    ws.cell(row=r, column=1).font = bold_navy
    r += 1

r += 2
# ═══════ SECTION 5: Collections Chain ═══════
ws.cell(row=r, column=1, value='5. COLLECTIONS CHAIN (Cash Flow Efficiency)').font = subtitle_font
r += 1
ws.cell(row=r, column=1, value='Collections % = Collections ÷ Revenue. Low % means high package redemption, gift cards, or AR leakage.').font = small_font
r += 1

coll_chains = [
    ['+1% collections rate', '+$0.94 collections per billable hour', 'Collections as % of revenue varies widely. AMP avg: 96.4%. Low % = heavy package/prepaid redemption.', '1. Collect payment at point of sale (never let balances walk out)\n2. Cherry/PatientFi financing for large treatment plans\n3. Follow up on outstanding balances within 7 days\n4. Review package redemption mix — too high = future revenue erosion\n5. Reduce gift card float (encourage redemption)\n6. Weekly AR aging review by PM', 'Each 1% improvement in coll % = +$0.94/billable hr in cash flow'],
    ['Promote Cherry financing', '+$500-2,000 per financed patient', 'Financing removes price objection → larger treatment plans → immediate cash collection (Cherry pays practice directly).', '1. Encourage application BEFORE appointment (resolve financing in advance)\n2. Position as budgeting tool, not financial need\n3. Present at consultation when plan is finalized\n4. Track application-to-treatment conversion rate\n5. Monthly Cherry volume targets per location\n6. Allē Cherry contribution to rebate tier progression', 'Higher financing = larger treatment plans + immediate cash + Allē rebate'],
    ['Allē coupon wallet check at every checkout', 'Prevents $50-200 negative patient experience', 'Checking rewards before payment prevents patient discovering missed points later → complaint → trust erosion.', '1. Ask patient to open Allē wallet before processing payment\n2. Check for manufacturer coupons, expiring points, gift cards\n3. Apply all available rewards proactively\n4. Post QR signage at checkout desk\n5. Script: "Let\'s check your rewards before we finish up"\n6. Track rewards application rate (target: 100%)', 'Prevents lost patients from bad checkout experience. Retention value: $2,000-5,000 LTV per saved patient.'],
]

for j, h in enumerate(headers):
    ws.cell(row=r, column=j+1, value=h)
hdr(ws, r, len(headers))
r += 1

for i, row_data in enumerate(coll_chains):
    for j, val in enumerate(row_data):
        ws.cell(row=r, column=j+1, value=val)
    drow(ws, r, len(headers), alt=(i%2==0))
    ws.cell(row=r, column=1).font = bold_navy
    r += 1

r += 2
# ═══════ SECTION 6: Cascade Summary ═══════
ws.cell(row=r, column=1, value='6. FULL CASCADE: HOW ONE ACTION RIPPLES THROUGH ALL KPIs').font = subtitle_font
r += 1
ws.cell(row=r, column=1, value='Example: Adding pre-consultation calls triggers a chain reaction across 7+ KPIs').font = small_font
r += 2

cascade_headers = ['Action', 'Step 1', 'Step 2', 'Step 3', 'Step 4', 'Net Revenue Impact']
for j, h in enumerate(cascade_headers):
    ws.cell(row=r, column=j+1, value=h)
hdr(ws, r, len(cascade_headers))
r += 1

cascades = [
    ['Implement pre-consultation calls', 'Show rate improves → cancel/no-show drops 2-3%', 'Unique patients increase +5-10/wk', 'Treatment acceptance rate rises → higher rev/patient', 'Rebooking rate increases (rapport built pre-visit)', '+$5,000-8,000/wk'],
    ['Full-face botox assessment training', 'Avg units increase from 35→50', 'NTX revenue rises +$5,954/wk per location', 'Patient satisfaction improves (better results)', 'Rebooking increases (patients see difference)', '+$6,000-10,000/wk'],
    ['Close morning shift gaps', 'Utilization jumps 40%→85%', 'Rev/billable hr nearly doubles', 'Same revenue in fewer paid hours = lower labor cost', 'Provider satisfaction improves (less idle time)', '+$3,000-5,000/wk saved in labor'],
    ['Activate Allē lapsed outreach', 'Reactivate 5-10 patients/month per location', 'Fill empty slots with zero acquisition cost', 'Returning patients tend to spend more (trust built)', 'Points redemption → positive checkout experience', '+$2,000-4,500/wk'],
    ['Cross-sell NTX+Filler combos', 'Dual treatment rate increases 10→25%', 'Rev/patient jumps ($555→$1,528 for combo visits)', 'Allergan rebate tier improves → lower COGS', 'Patient gets better outcome → higher retention', '+$4,000-8,000/wk'],
    ['Eliminate wasteful Thursday shifts', 'Remove 2-3 shifts with <2 booked hrs', 'Save 16-24 scheduled hours of labor cost', 'Utilization rate improves (denominator shrinks)', 'Rev/scheduled hr increases significantly', '+$2,000-4,000/wk saved'],
]

for i, row_data in enumerate(cascades):
    for j, val in enumerate(row_data):
        ws.cell(row=r, column=j+1, value=val)
    drow(ws, r, len(cascade_headers), alt=(i%2==0))
    ws.cell(row=r, column=1).font = bold_navy
    ws.cell(row=r, column=6).font = Font(name='Arial', bold=True, color=GREEN, size=10)
    ws.cell(row=r, column=6).fill = green_bg
    r += 1

r += 2
ws.cell(row=r, column=1, value='THE PRIORITY FORMULA: Focus on Patient Volume first (biggest impact), then Revenue Per Patient (second biggest), then Efficiency (cost containment). Volume × Yield = Revenue. Efficiency = Profitability.').font = Font(name='Arial', bold=True, italic=True, color=NAVY, size=11)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)

# Column widths
for i, w in enumerate([30, 28, 42, 55, 32, 28]):
    ws.column_dimensions[get_column_letter(i+1)].width = w

wb.save('AMP_KPI_Operational_Playbook_v3.xlsx')
print('Saved with new Metric Chains sheet')
print('Final sheets:', wb.sheetnames)
